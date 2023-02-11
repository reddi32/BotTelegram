import openpyxl
import requests

# Apri il file Excel
wb = openpyxl.load_workbook('urls.xlsx')

# Seleziona la prima scheda
sheet = wb.active

# Cicla attraverso le righe della colonna A
for row in sheet.iter_rows(values_only=True, min_row=1, max_col=1, max_row=sheet.max_row):
    url = row[0]
    # Inviami la richiesta di rimozione di indicizzazione a Google
    requests.get("https://www.google.com/search?q=cache:" + url)
